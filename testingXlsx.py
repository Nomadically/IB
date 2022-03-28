import csv
import subprocess
import time
import pyperclip
from ahk import AHK
import keyboard
import datetime

from openpyxl import load_workbook
wb = load_workbook(filename='Manifest-EmailTemplate.xlsx')
#worksheet = wb.get_sheet_by_name('Sheet1')
sheet1 = wb['Sheet1']
#worksheet['D3']='Whatever you want to put in D3'
sheet1['D3'] = 'Testing 03/27/2022'
wb.save('Manifest-EmailTemplate-Test.xlsx')

prisons = {}

place = 'Valley'

with open('CAprisonsCSV.csv') as csv_file:
    # reading the csv file using DictReader
    csv_reader = csv.DictReader(csv_file)
    for row in csv_reader:
        if place in row['prison']:
            print(row)
            # break

date = 'Mar 12th-Apr 1st 2022'
tbirdPath = r'E:\Program Files\Mozilla Thunderbird\thunderbird.exe'
to = 'ykchaudry@gmail.com'
attachment = r'C:\Users\think\Documents\progressQuestion.png'
subject = """Master List for IslamicBookstore.com -"""+date
body = ''
message = """asalaamu'alaikum (Peace be with you),

Greetings,

Please find attached the master list for the package(s) being shipped this week.

It is expected to arrive within 10 to 14 business days via USPS or FedEx.

IF YOU DO NOT RECEIVE THESE PACKAGES, PLEASE NOTIFY US AS SOON AS POSSIBLE.


If you have any questions or concerns, please let us know.

Thank you, 
"""
pyperclip.copy(message)
composeCommand = 'format=html,to={},subject={},body={},attachment={}'.format(to, subject, body, attachment)
subprocess.Popen([tbirdPath, '-compose', composeCommand])
time.sleep(1)
keyboard.send('ctrl+v')
# keyboard.write(message) -- this also works but not needed, faster to paste

#keyboard.send('ctrl+enter') # 3/4/22: get this to work!--works

print('worked till here')

us_state_abbrev = {
    'Alabama': 'AL',
    'Alaska': 'AK',
    'American Samoa': 'AS',
    'Arizona': 'AZ',
    'Arkansas': 'AR',
    'California': 'CA',
    'Colorado': 'CO',
    'Connecticut': 'CT',
    'Delaware': 'DE',
    'District of Columbia': 'DC',
    'Florida': 'FL',
    'Georgia': 'GA',
    'Guam': 'GU',
    'Hawaii': 'HI',
    'Idaho': 'ID',
    'Illinois': 'IL',
    'Indiana': 'IN',
    'Iowa': 'IA',
    'Kansas': 'KS',
    'Kentucky': 'KY',
    'Louisiana': 'LA',
    'Maine': 'ME',
    'Maryland': 'MD',
    'Massachusetts': 'MA',
    'Michigan': 'MI',
    'Minnesota': 'MN',
    'Mississippi': 'MS',
    'Missouri': 'MO',
    'Montana': 'MT',
    'Nebraska': 'NE',
    'Nevada': 'NV',
    'New Hampshire': 'NH',
    'New Jersey': 'NJ',
    'New Mexico': 'NM',
    'New York': 'NY',
    'North Carolina': 'NC',
    'North Dakota': 'ND',
    'Northern Mariana Islands':'MP',
    'Ohio': 'OH',
    'Oklahoma': 'OK',
    'Oregon': 'OR',
    'Pennsylvania': 'PA',
    'Puerto Rico': 'PR',
    'Rhode Island': 'RI',
    'South Carolina': 'SC',
    'South Dakota': 'SD',
    'Tennessee': 'TN',
    'Texas': 'TX',
    'Utah': 'UT',
    'Vermont': 'VT',
    'Virgin Islands': 'VI',
    'Virginia': 'VA',
    'Washington': 'WA',
    'West Virginia': 'WV',
    'Wisconsin': 'WI',
    'Wyoming': 'WY'
}
# creating nested dictionary of each facility, map to inmate-manifest
send_to_prisons = {
    'facility 1': {'manifest': '54321', 'nameID': 'Joe Blow 12345'},
    'facility 2': {'manifest': 'aa', 'nameID': 'Joe Cool 98765'}
}

print(us_state_abbrev['Virginia'])

testprison = 'facility 2'

if send_to_prisons[testprison] is not None:
    print('good')
    print(send_to_prisons[testprison]['manifest'])



prisons = {
'Avenal State Prison': {'manifest':'ref#', 'nameID':''},
'California City Correctional Facility': {'manifest':'ref#', 'nameID':''},
'Calipatria State Prison': {'manifest':'ref#', 'nameID':''},
'California Correctional Center': {'manifest':'ref#', 'nameID':''},
'California Correctional Institution': {'manifest':'ref#', 'nameID':''},
'California Correctional Women’s Facility': {'manifest':'ref#', 'nameID':''},
'California State Prison, Centinela': {'manifest':'ref#', 'nameID':''},
'California Health Care Facility': {'manifest':'ref#', 'nameID':''},
'California Institution for Men': {'manifest':'ref#', 'nameID':''},
'California Institution for Women': {'manifest':'ref#', 'nameID':''},
'California Men’s Colony': {'manifest':'ref#', 'nameID':''},
'California Medical Facility': {'manifest':'ref#', 'nameID':''},
'CSP, Corcoran': {'manifest':'ref#', 'nameID':''},
'California Rehabilitation Center': {'manifest':'ref#', 'nameID':''},
'Correctional Training Facility': {'manifest':'ref#', 'nameID':''},
'Chuckawalla Valley State Prison': {'manifest':'ref#', 'nameID':''},
'Folsom State Prison': {'manifest':'ref#', 'nameID':''},
'High Desert State Prison': {'manifest':'ref#', 'nameID':''},
'Ironwood State Prison': {'manifest':'ref#', 'nameID':''},
'Kern Valley State Prison': {'manifest':'ref#', 'nameID':''},
'CSP, Los Angeles County': {'manifest':'ref#', 'nameID':''},
'Mule Creek State Prison': {'manifest':'ref#', 'nameID':''},
'North Kern State Prison': {'manifest':'ref#', 'nameID':''},
'Pelican Bay State Prison': {'manifest':'ref#', 'nameID':''},
'Pleasant Valley State Prison': {'manifest':'ref#', 'nameID':''},
'R. J. Donovan Correctional Facility': {'manifest':'ref#', 'nameID':''},
'California State Prison, Sacramento': {'manifest':'ref#', 'nameID':''},
'SATF SP, Corcoran': {'manifest':'ref#', 'nameID':''},
'Sierra Conservation Center': {'manifest':'ref#', 'nameID':''},
'California State Prison, Solano': {'manifest':'ref#', 'nameID':''},
'San Quentin State Prison': {'manifest':'ref#', 'nameID':''},
'Salinas Valley State Prison': {'manifest':'ref#', 'nameID':''},
'Valley State Prison': {'manifest':'ref#', 'nameID':''}
}


testprison2 = 'Salinas Valley State Prison'


for pr, pi in prisons.items():
    print(pr)

    for key in pi:
        print(key+':', pi[key])

print('ok till here 2 for now')

testprison3 = 'Salinas Valley State Prison'

if prisons[testprison3] is not None and prisons[testprison3]['nameID'] != '':
    print(prisons[testprison3]['manifest'])
else:
    print('need a name here')


"""
3/27/22:

right now, can create email with defined fields, and add attachment
now to send, done






"""





# if send_to_prisons[testprison] is not None:
#     print('good')
#     print(send_to_prisons[testprison]['manifest'])


# f = open('CAprisonsCSV.csv', 'rt')
# csv_reader = csv.DictReader(f, escapechar='\\')
# for row in csv_reader:
#     prisons.append(row)

# dicted = dict(list(csv_reader[0]))

# print(dicted)


# opening the csv file
# with open('CAprisonsCSV.csv') as csv_file:
#     # reading the csv file using DictReader
#     csv_reader = csv.DictReader(csv_file)
#
#     # converting the file to dictionary
#     # by first converting to list
#     # and then converting the list to dict
#     dict_from_csv = dict(list(csv_reader)[1])
#
#     # making a list from the keys of the dict
#     list_of_column_names = list(dict_from_csv.keys())
#
#     # displaying the list of column names
#     print("List of column names : ",
#           list_of_column_names)
#     print(dict_from_csv)



# with open('CAprisonsCSV.csv', mode='r') as infile:
#     reader = csv.reader(infile)
#     with open('CAprisonsCSV-new.csv', mode='w') as outfile:
#         writer = csv.writer(outfile)
#         mydict = {rows[0]:rows[1] for rows in reader}

# opening the csv file by specifying
# the location
# with the variable name as csv_file
# with open('CAprisonsCSV.csv') as csv_file:
#     # creating an object of csv reader
#     # with the delimiter as ,
#     csv_reader = csv.reader(csv_file, delimiter=',')
#
#     # list to store the names of columns
#     list_of_column_names = []
#
#     # loop to iterate through the rows of csv
#     for row in csv_reader:
#         # adding the first row
#         list_of_column_names.append(row)
#
#         # breaking the loop after the
#         # first iteration itself
#         break
#
# # printing the result
# print("List of column names : ",
#       list_of_column_names[0])







# wb = load_workbook(filename = 'Test-Export-CAprisons.xlsx')
# sheet_ranges = wb['Customer #']
# print(sheet_ranges['A9'].value)

#this function finds data from CSV based on facility, so facility's address info
def fieldentry2(prison):
    #f = open('customers10.csv', 'rt')
    f = open('CAprisonsCSV.csv', 'rt')
    csv_reader = csv.DictReader(f, escapechar='\\')
    for row in csv_reader:
        if prison in row['prison']: # >>>>>>>>>>>>>>>>> need to update this logic, add 2nd check
            n = row['prison']         # like matching  as well as prison !! <<<<--- 09/10/21
            email = row['email']
            # if row['street2'] is not None:
            #     strt2 = row['street2'], #need to add this also, 12/01/21, done on 12/09/21
            # zpp = row['zip']
            break
    f.close
    print(row)
    if row['prison'] is not None and len(row['email']) > 1:
        # 01/28/22: fix the below, the str casting of str2, looks like a tuple inside mailware
        results = [n.strip(), email.strip()] #add str2 to this array <<< 12/01/21, then continue tracing where else code needs to be updated
    else:
        results = [n.strip(), email.strip()]
    return results

#fieldentry2('Salinas')

# class HashTable:
#
#     # Create empty bucket list of given size
#     def __init__(self, size):
#         self.size = size
#         self.hash_table = self.create_buckets()
#
#     def create_buckets(self):
#         return [[] for _ in range(self.size)]
#
#     # Insert values into hash map
#     def set_val(self, key, val):
#
#         # Get the index from the key
#         # using hash function
#         hashed_key = hash(key) % self.size
#
#         # Get the bucket corresponding to index
#         bucket = self.hash_table[hashed_key]
#
#         found_key = False
#         for index, record in enumerate(bucket):
#             record_key, record_val = record
#
#             # check if the bucket has same key as
#             # the key to be inserted
#             if record_key == key:
#                 found_key = True
#                 break
#
#         # If the bucket has same key as the key to be inserted,
#         # Update the key value
#         # Otherwise append the new key-value pair to the bucket
#         if found_key:
#             bucket[index] = (key, val)
#         else:
#             bucket.append((key, val))
#
#     # Return searched value with specific key
#     def get_val(self, key):
#
#         # Get the index from the key using
#         # hash function
#         hashed_key = hash(key) % self.size
#
#         # Get the bucket corresponding to index
#         bucket = self.hash_table[hashed_key]
#
#         found_key = False
#         for index, record in enumerate(bucket):
#             record_key, record_val = record
#
#             # check if the bucket has same key as
#             # the key being searched
#             if record_key == key:
#                 found_key = True
#                 break
#
#         # If the bucket has same key as the key being searched,
#         # Return the value found
#         # Otherwise indicate there was no record found
#         if found_key:
#             return record_val
#         else:
#             return "No record found"
#
#     # Remove a value with specific key
#     def delete_val(self, key):
#
#         # Get the index from the key using
#         # hash function
#         hashed_key = hash(key) % self.size
#
#         # Get the bucket corresponding to index
#         bucket = self.hash_table[hashed_key]
#
#         found_key = False
#         for index, record in enumerate(bucket):
#             record_key, record_val = record
#
#             # check if the bucket has same key as
#             # the key to be deleted
#             if record_key == key:
#                 found_key = True
#                 break
#         if found_key:
#             bucket.pop(index)
#         return
#
#     # To print the items of hash map
#     def __str__(self):
#         return "".join(str(item) for item in self.hash_table)
#
#
# hash_table = HashTable(50)
#
# # insert some values
# hash_table.set_val('gfg@example.com', 'some value')
# print(hash_table)
# print()
#
# hash_table.set_val('portal@example.com', 'some other value')
# print(hash_table)
# print()
#
# # search/access a record with key
# print(hash_table.get_val('portal@example.com'))
# print()
#
# # delete or remove a value
# hash_table.delete_val('portal@example.com')
# print(hash_table)



