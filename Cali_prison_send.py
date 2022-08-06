

import csv
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
import PySimpleGUI as sg
import keyboard
import pyperclip
from openpyxl import load_workbook

sg.theme('DarkAmber')  # Add a touch of color
# All the stuff inside your window.

prisons = {
    'Avenal State Prison': {'email'},
    'California City Correctional Facility': {'email'},
    'Calipatria State Prison': {'email'},
    'California Correctional Center': {'email'},
    'California Correctional Institution': {'email'},
    "California Correctional Women's Facility": {'email'},
    'California State Prison, Centinela': {'email'},
    'California Health Care Facility': {'email'},
    'California Institution for Men': {'email'},
    'California Institution for Women': {'email'},
    "California Men's Colony": {'email'},
    'California Medical Facility': {'email'},
    'California State Prison, Corcoran': {'email'},
    'California Rehabilitation Center': {'email'},
    'Correctional Training Facility': {'email'},
    'Chuckawalla Valley State Prison': {'email'},
    'Folsom State Prison': {'email'},
    'High Desert State Prison': {'email'},
    'Ironwood State Prison': {'email'},
    'Kern Valley State Prison': {'email'},
    'California State Prison, Los Angeles County': {'email'},
    'Mule Creek State Prison': {'email'},
    'North Kern State Prison': {'email'},
    'Pelican Bay State Prison': {'email'},
    'Pleasant Valley State Prison': {'email'},
    'Richard J. Donovan Correctional Facility': {'email'},
    'California State Prison, Sacramento': {'email'},
    'SATF California State Prison, Corcoran': {'email'},
    'Sierra Conservation Center': {'email'},
    'California State Prison, Solano': {'email'},
    'San Quentin State Prison': {'email'},
    'Salinas Valley State Prison': {'email'},
    'Valley State Prison': {'email'},
    'Wasco State Prison': {'email'}
}

options = {
    "font": ('Helvetica', 16),
    "size": (45, 5),
    "readonly": True,
    "enable_events": True,
    # "element_justification": 'center'
}
inmate_name = sg.Input(key='name1', tooltip='Enter inmate first name here = ', size=(30,1))
id_num = sg.Input(key='id_1', tooltip='Enter inmate number here =: ', size=(30,1))

dir_process = sg.Button('Create directory first')

dat1 = sg.CalendarButton(button_text='Select From date:', target='date1', format='%m-%d-%Y', key='d1')
dat2 = sg.CalendarButton(button_text='Select To date:', target='date2', format='%m-%d-%Y', key='d2')
# z = sg.Button('Clear All Data')
layout = [[sg.Text('Program to auto-compile/send California Manifests')],
         # [sg.Text('Enter inmate name:'), inmate_name],
          #[sg.Text('Enter inmate ID:'), id_num],
          #[dir_process],
          #sg.Multiline(s=(25, 2))], #this is value[0]
          [sg.Text('Enter info as follows = inmate name,inmate ID,invoice number; if 2+ inmates, separate with a semicolon.')],
          [sg.Multiline(s=(25, 2))],
          [dat1, sg.InputText('   ', key='date1', size=(10, 1), tooltip='<- Click button to choose first date.')],
          [dat2, sg.InputText('   ', key='date2', size=(10, 1), tooltip='<- Click button to choose second date.')],
          [sg.Combo(sorted(list(prisons.keys())), **options, key="mainproject",
                    tooltip='<-Choose facility, state prisons of California CDCR->')],
          [sg.Button('Process Prison')],
          # z,
          [sg.Button('Ok'), sg.Button('Cancel')]]


# Create the Window
# below will help to correct dir location at each run, need to finish implementing, 4/22/22
# current_location = input("What is current folder directory for this week?")
# arg1 = sys.argv[1]

window = sg.Window('Auto-Email California Prison Manifests', layout)
# Event Loop to process "events" and get the "values" of the inputs

#needed_path = input("what is file path for this week?\n")

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':  # if user closes window or clicks cancel
        break
    # if event == 'Clear All Data': #implement this later, to clear fields
    #     window['fn'].update('')
    #     window['ln'].update('')
    #     window['inm'].update('')
    #     window['mainproject'].update('')
    #     window['subproject'].update('')
    #     keyboard.send('tab')
    if event == 'Process Prison':
        print(values[0])
        if values['date1'] == '' or values['date2'] == '' or values['mainproject'] == '' or values[0] == '':
            sg.popup("Please enter all required data first!", title="Error")
            print(values['date1'], values['date2'], values['mainproject'], values[0])
            continue
        else:
            with open('CAprisonsCSV.csv') as csv_file:
                # reading the csv file using1 DictReader
                csv_reader = csv.DictReader(csv_file)
                for row in csv_reader:
                    if values['mainproject'] == row['prison']:
                        print(row['acronym'])
                        # myDict["niche"] = "programming"
                        # inm_info = {values[0]: values[1]}
                        # print('You entered ', values[0], values[1])

                        date1 = values['date1']
                        date2 = values['date2']
                        # print('this is what this created dict looks like:')
                        # print(inm_info)
                        print("at line 128")
                        print(date1)
                        print(date2)
                        # wb = load_workbook(filename='Manifest-EmailTemplate.xlsx')
                        wb = load_workbook(filename='Manifest-EmailTemplate-Test-2.xlsx')
                        # worksheet = wb.get_sheet_by_name('Sheet1')
                        sheet1 = wb['Sheet1']
                        # worksheet['D3']='Whatever you want to put in D3'
                        sheet1['A2'] = values['mainproject']  # <<<--- facility name goes in here
                        sheet1['B5'] = 'From ' + date1 + ' to ' + date2  # <<<--- date ranges go in here
                        sheet1['B6'] = row['email']  # <<<--- prison's email goes in here
                        # sheet1['A11'] = values[1]  # <<<--- manifest #s goes in here
                        # sheet1['B11'] = values[0]  # <<<--- inmates name go in here
                        # sheet1['C11'] = values[0]  # <<<--- inmates ID go in here
                        if ';' in values[0]:
                            inmate_group = values[0].split(';')
                            total = len(inmate_group)
                            count = 0
                            for n in inmate_group:
                                inmate = inmate_group[count].split(',')
                                inmate_name = inmate[0].strip()
                                inmate_id = inmate[1].strip()
                                manifest_ref = inmate[2].strip()
                                sheet1['A1'+str(count)] = manifest_ref
                                sheet1['B1'+str(count)] = inmate_name
                                sheet1['C1'+str(count)] = inmate_id
                                count += 1
                        else:
                            inmate = values[0].split(',')
                            inmate_name = inmate[0].strip()
                            inmate_id = inmate[1].strip()
                            manifest_ref = inmate[2].strip()
                            sheet1['A10'] = manifest_ref  # <<<--- manifest #s goes in here
                            sheet1['B10'] = inmate_name  # <<<--- inmates name go in here
                            sheet1['C10'] = inmate_id  # <<<--- inmates ID go in here
                        time.sleep(1)
                        filename = 'Manifest-Email-' + row['acronym']
                        # file = "shortcut_folder/filename"
                        # os.path.abspath(file) = "C:/Desktop/shortcut_folder/filename"
                        # fname = os.path.join(my_dir, file_name)
                        fpath_parent = 'C:\\Users\\ib\\Documents\\CA-Manifests\\'
                        weekly_dir = date1+"-to-"+date2
                        if len(weekly_dir) < 7:
                            sg.popup("Dates not entered, please enter both From and To.")
                            continue
                        full_path = os.path.join(fpath_parent, weekly_dir)
                        try:
                            os.mkdir(full_path)
                            time.sleep(1)
                        except FileExistsError:
                            sg.popup("Double check name, date range may already exist.")
                            window['Create directory first'].update(disabled=False)
                            continue
                        fname = os.path.join(full_path, filename + '.xlsx')
                        wb.save(fname)
                        time.sleep(1)
                        #date = date1 + '-' + date2
                        tbirdPath = r'C:\Program Files\Mozilla Thunderbird\thunderbird.exe'
                        #to = 'ykchaudry@gmail.com'

                        #cc = 'ykchaudry@gmail.com, ykchaudry@gmail.com' #change emails to prison email+AK later
                        p = Path(fname).resolve()
                        print(p)
                        # attachment = r'C:\Users\think\Documents\progressQuestion.png'
                        attachment = p

                        os.startfile(attachment)
                        time.sleep(2)
                        keyboard.send('alt+f+e')
                        keyboard.send('enter')
                        keyboard.send('enter')
                        time.sleep(1)
                        filetitle = 'Manifest-Email-'+row['acronym']
                        keyboard.write(filetitle)
                        time.sleep(1)
                        keyboard.send('enter')
                        time.sleep(2)
                        keyboard.send('ctrl+q')
                        #keyboard.send('alt+s')
                        #time.sleep(0.5)
                        #keyboard.send('ctrl+q')
                        time.sleep(1)
                        fname2 = os.path.join(full_path, filename + '.pdf')
                        p2 = Path(fname2).resolve()
                        attachment2 = p2
                        # don't forget to modify this line below ?>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> !!!<<<<<
                        #no more need for this as of 4/15/22
                        #date = 'Mar 12th-Apr 1st 2022' ----update as needed
                        #date = ''
                        # date = ''
                        # date = ''
                        # date = ''
                        # date = ''
                        # date = ''
                        # date = ''
                        #date = 'April 8th to April 14th 2022'
                        date = f'{date1} to {date2}'
                        subject = """Manifest Master List-IslamicBookstore.com- """ + date
                        body = ''
                        #to = 'yousaf@islamicbookstore.com'
                        # if row['email'].split(',') > 1:
                        #     to =
                        to = row['email']
                        print(row['email'])
                        cc = 'adnank@islamicbookstore.com'
                        message = """Greetings,
    
Please find attached the manifest master list for the package(s) being shipped this week.
   
It is expected to arrive within 10 to 14 business days via USPS or FedEx/UPS.
    
IF YOU DO NOT RECEIVE THESE PACKAGES, PLEASE NOTIFY US AS SOON AS POSSIBLE.
    
    
If you have any questions or concerns, please let us know.
    
Thank you, 
Yousaf 
    
Inventory & Sales, http://IslamicBookstore.com
Metric Networks Inc. DBA http://IslamicBookstore.com
3918 Vero Road, Suite I. ("i" as in 'iris') --- Baltimore, MD 21227 USA
410.675.0040 x16 (direct line) 
                        """
                        pyperclip.copy(message)
                        # add cc, later -- cc={},
                        composeCommand = 'format=html,to={},subject={},cc={},body={},attachment={}'.format(to,subject,cc,body,attachment2)
                        subprocess.Popen([tbirdPath, '-compose', composeCommand])
                        print('good till before sending email')
                        time.sleep(1)

                        #time.sleep(1)
                        keyboard.send('ctrl+v')
                        time.sleep(1)
                        # keyboard.write(message) -- this also works but not needed, faster to paste
                        keyboard.send('alt+r')
                        time.sleep(0.5)
                        keyboard.send('down')
                        time.sleep(2)
                        keyboard.send('tab')
                        keyboard.write('lisa@islamicbookstore.com')
                        #keyboard.send('ctrl+enter') # 3/4/22: get this to work!-----works
                        print("good till almost send email")

                        break
    #  and values['date1'] != '' and values['date2'] not in ['', None]
    if event == 'Create directory first':
        if values['d1'] == '' or values['d2'] == '':
            sg.popup("Please first input start date and end date.", title="Error")
            continue
        else:
            try:
                window['Create directory first'].update(disabled=True)
                date1 = values['date1']
                date2 = values['date2']
                fpath_parent = 'C:\\Users\\ib\\Documents\\CA-Manifests\\'
                weekly_dir = date1 + "-to-" + date2
                full_path = os.path.join(fpath_parent, weekly_dir)
                os.mkdir(full_path)
                time.sleep(1)
                # fpath_weekly = 'C:\\Users\\ib\\Documents\\CA-Manifests\\Manifest-Weekly.xlsx'
                # wb2 = load_workbook(fpath_weekly)
                # os.startfile(fpath_weekly)
                sg.popup("Directory created, now proceed to Excel file and enter inmate info; once done, click Process Prison.")
                continue
            except FileExistsError:
                sg.popup("Double check name, date range may already exist.")
                window['Create directory first'].update(disabled=False)
                continue

window.close()
