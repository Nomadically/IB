
from openpyxl import load_workbook

wb = load_workbook(filename='Manifest-Weekly.xlsx')

sheet = wb['Sheet3']
# worksheet['D3']='Whatever you want to put in D3'
#sheet['A2'] = values['mainproject']

#print(sheet['C2'].value)

names = sheet['C2'].value
current = 'Avenal State Prison'
#as of 8/5/22 = these can access any rows and columns in an excel file xlsx



# first need to match facility in mainproject = to prison in manifest-weekly file
# then need to go to inmates row, access values and str.split them according to ;

#inmates_test = 'john doe, A123456, 433112; johny doey, A1245453456, 4335112; johnee doeee, A123456766, 43315512 '
#inmates_test = 'john doe,A123456,433112'

inmates_test = names
if ';' in inmates_test:
    new_dict = inmates_test.split(';')
    print(new_dict[0])
    print(new_dict[1])
else:
    single = inmates_test.split(',')
    print(single[0].title())
    print(single[1])
    print(single[2])

stored = {1: names}

#print(stored)

test = {'fname last name', '123455', 445311}
test2 =['fname last name', '123455', 445311]

for col in sheet['A']:
    print(col.value)
    if col.value == current:
        print('this is where we found it')

for row in sheet[1]:
    print(row.value)


prisons = []
for col in sheet['A']:
    prisons.append(col.value)
    #print('test')

t = 0

if current in prisons:
    print('finally?')
    t = prisons.index(current)
print(t)
# ws.cell(row, column)
print(sheet[('C'+str(t))].value)